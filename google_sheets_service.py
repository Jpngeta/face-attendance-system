"""
Google Sheets Service for Face Attendance System
Creates and populates Google Sheets with attendance data
"""
import os
from datetime import datetime
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class GoogleSheetsService:
    """Service for creating and managing Google Sheets with attendance data"""

    SCOPES = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive.file'  # Narrow scope - only files created by this app
    ]

    def __init__(self, credentials_path=None):
        """
        Initialize Google Sheets service

        Args:
            credentials_path: Path to service account credentials JSON file
        """
        # 1. Try the argument passed in
        # 2. Try the Environment Variable
        if credentials_path is None:
            credentials_path = os.getenv('GOOGLE_CREDENTIALS_PATH')

        # 3. Fallback: Check for the standard file name in the current directory
        if not credentials_path and os.path.exists('google-credentials.json'):
            credentials_path = 'google-credentials.json'
            print(f"[INFO] Auto-detected credentials: {credentials_path}")

        # 4. Fallback: Check for absolute path (common in your setup)
        if not credentials_path and os.path.exists('/home/jpngeta/face-attendance-system/google-credentials.json'):
            credentials_path = '/home/jpngeta/face-attendance-system/google-credentials.json'
            print(f"[INFO] Auto-detected credentials at absolute path: {credentials_path}")

        if not credentials_path or not os.path.exists(credentials_path):
            raise ValueError(f"Google credentials not found at: {credentials_path}")

        self.credentials = service_account.Credentials.from_service_account_file(
            credentials_path,
            scopes=self.SCOPES
        )

        self.service = build('sheets', 'v4', credentials=self.credentials)
        self.drive_service = build('drive', 'v3', credentials=self.credentials)

    def create_attendance_sheet(self, session_data, attendance_records):
        """
        Create a new Google Sheet with attendance data

        Args:
            session_data: Dictionary containing session information
            attendance_records: List of attendance record dictionaries

        Returns:
            dict: {'spreadsheet_id': str, 'spreadsheet_url': str}

        Raises:
            HttpError: If Google Sheets creation fails (caller should use Excel fallback)
        """
        # Generate sheet title
        session_name = session_data.get('session_name', 'Attendance')
        date_str = datetime.now().strftime('%Y-%m-%d')
        sheet_title = f"Attendance - {session_name} - {date_str}"

        # Create spreadsheet
        spreadsheet = {
            'properties': {
                'title': sheet_title
            },
            'sheets': [{
                'properties': {
                    'title': 'Attendance',
                    'gridProperties': {
                        'frozenRowCount': 1  # Freeze header row
                    }
                }
            }]
        }

        result = self.service.spreadsheets().create(body=spreadsheet).execute()
        spreadsheet_id = result['spreadsheetId']
        spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"

        print(f"[INFO] Created Google Sheet: {sheet_title}")
        print(f"[INFO] URL: {spreadsheet_url}")

        # Share the sheet with instructor if email is provided
        instructor_email = session_data.get('instructor_email')
        if instructor_email:
            self._share_sheet(spreadsheet_id, instructor_email)

        # Populate the sheet with data
        self._populate_sheet(spreadsheet_id, session_data, attendance_records)

        return {
            'spreadsheet_id': spreadsheet_id,
            'spreadsheet_url': spreadsheet_url,
            'title': sheet_title
        }

    def _share_sheet(self, spreadsheet_id, email_address):
        """
        Share the Google Sheet with a specific email address

        Args:
            spreadsheet_id: ID of the spreadsheet
            email_address: Email address to share with
        """
        try:
            # Create permission for the user
            permission = {
                'type': 'user',
                'role': 'writer',  # Can edit the sheet
                'emailAddress': email_address
            }

            self.drive_service.permissions().create(
                fileId=spreadsheet_id,
                body=permission,
                sendNotificationEmail=False  # Don't send email notification
            ).execute()

            print(f"[INFO] Shared sheet with {email_address}")

        except HttpError as error:
            print(f"[WARNING] Failed to share sheet with {email_address}: {error}")
            # Non-critical, don't raise

    def _populate_sheet(self, spreadsheet_id, session_data, attendance_records):
        """
        Populate the Google Sheet with session info and attendance data

        Args:
            spreadsheet_id: ID of the spreadsheet
            session_data: Session information
            attendance_records: List of attendance records
        """
        try:
            # Prepare header and data
            values = self._prepare_sheet_data(session_data, attendance_records)

            # Write data to sheet
            body = {
                'values': values
            }

            self.service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range='Attendance!A1',
                valueInputOption='RAW',
                body=body
            ).execute()

            # Format the sheet
            self._format_sheet(spreadsheet_id)

            print(f"[INFO] Populated sheet with {len(attendance_records)} attendance records")

        except HttpError as error:
            print(f"[ERROR] Failed to populate sheet: {error}")
            raise

    def _prepare_sheet_data(self, session_data, attendance_records):
        """
        Prepare data for the Google Sheet

        Args:
            session_data: Session information
            attendance_records: List of attendance records

        Returns:
            list: 2D list of values for the sheet
        """
        # Session info section
        values = [
            ['ATTENDANCE REPORT'],
            [],
            ['Session:', session_data.get('session_name', 'N/A')],
            ['Course Code:', session_data.get('course_code', 'N/A')],
            ['Course Name:', session_data.get('course_name', 'N/A')],
            ['Instructor:', session_data.get('instructor_name', 'N/A')],
            ['Location:', session_data.get('location', 'N/A')],
            ['Start Time:', session_data.get('start_time', 'N/A')],
            ['End Time:', session_data.get('end_time', 'N/A')],
            ['Total Students:', str(len(attendance_records))],
            [],
            # Header row for attendance data
            ['Student ID', 'Name', 'Email', 'Program', 'Time Marked', 'Status', 'Confidence Score']
        ]

        # Attendance records
        for record in attendance_records:
            values.append([
                record.get('student_student_id', 'N/A'),
                record.get('student_name', 'N/A'),
                record.get('student_email', 'N/A'),
                record.get('student_program', 'N/A'),
                record.get('timestamp', 'N/A'),
                record.get('status', 'present').upper(),
                f"{record.get('confidence_score', 0):.4f}" if record.get('confidence_score') else 'N/A'
            ])

        return values

    def _format_sheet(self, spreadsheet_id):
        """
        Apply formatting to the Google Sheet

        Args:
            spreadsheet_id: ID of the spreadsheet
        """
        try:
            requests = [
                # Bold title row
                {
                    'repeatCell': {
                        'range': {
                            'sheetId': 0,
                            'startRowIndex': 0,
                            'endRowIndex': 1
                        },
                        'cell': {
                            'userEnteredFormat': {
                                'textFormat': {
                                    'bold': True,
                                    'fontSize': 14
                                },
                                'horizontalAlignment': 'CENTER'
                            }
                        },
                        'fields': 'userEnteredFormat(textFormat,horizontalAlignment)'
                    }
                },
                # Bold header row (row 12, index 11)
                {
                    'repeatCell': {
                        'range': {
                            'sheetId': 0,
                            'startRowIndex': 11,
                            'endRowIndex': 12
                        },
                        'cell': {
                            'userEnteredFormat': {
                                'textFormat': {
                                    'bold': True
                                },
                                'backgroundColor': {
                                    'red': 0.8,
                                    'green': 0.8,
                                    'blue': 0.8
                                }
                            }
                        },
                        'fields': 'userEnteredFormat(textFormat,backgroundColor)'
                    }
                },
                # Auto-resize columns
                {
                    'autoResizeDimensions': {
                        'dimensions': {
                            'sheetId': 0,
                            'dimension': 'COLUMNS',
                            'startIndex': 0,
                            'endIndex': 7
                        }
                    }
                },
                # Freeze header row (row 12)
                {
                    'updateSheetProperties': {
                        'properties': {
                            'sheetId': 0,
                            'gridProperties': {
                                'frozenRowCount': 12
                            }
                        },
                        'fields': 'gridProperties.frozenRowCount'
                    }
                }
            ]

            body = {
                'requests': requests
            }

            self.service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body=body
            ).execute()

            print("[INFO] Applied formatting to sheet")

        except HttpError as error:
            print(f"[ERROR] Failed to format sheet: {error}")
            # Non-critical, so we don't raise

    def export_to_excel(self, spreadsheet_id, output_path):
        """
        Export Google Sheet to Excel file

        Args:
            spreadsheet_id: ID of the spreadsheet
            output_path: Path where Excel file should be saved

        Returns:
            str: Path to the created Excel file
        """
        try:
            # Read data from Google Sheet
            result = self.service.spreadsheets().values().get(
                spreadsheetId=spreadsheet_id,
                range='Attendance!A1:G1000'  # Adjust range as needed
            ).execute()

            values = result.get('values', [])

            if not values:
                raise ValueError("No data found in sheet")

            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Attendance'

            # Write data to Excel
            for row_idx, row in enumerate(values, start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)

                    # Format title row
                    if row_idx == 1:
                        cell.font = Font(bold=True, size=14)
                        cell.alignment = Alignment(horizontal='center')

                    # Format header row (row 12)
                    elif row_idx == 12:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center')

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save Excel file
            wb.save(output_path)
            print(f"[INFO] Exported to Excel: {output_path}")

            return output_path

        except HttpError as error:
            print(f"[ERROR] Failed to export to Excel: {error}")
            raise
        except Exception as error:
            print(f"[ERROR] Failed to create Excel file: {error}")
            raise


def create_and_export_attendance_report(session_data, attendance_records, output_dir='/tmp'):
    """
    Helper function to create Google Sheet and export to Excel

    Args:
        session_data: Dictionary with session information
        attendance_records: List of attendance record dictionaries
        output_dir: Directory where Excel file should be saved

    Returns:
        dict: {
            'spreadsheet_id': str,
            'spreadsheet_url': str,
            'excel_path': str,
            'title': str
        }
    """
    # Initialize service
    sheets_service = GoogleSheetsService()

    # Create Google Sheet
    sheet_result = sheets_service.create_attendance_sheet(session_data, attendance_records)

    # Generate Excel filename
    session_name = session_data.get('session_name', 'Attendance').replace(' ', '_')
    date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_filename = f"Attendance_{session_name}_{date_str}.xlsx"
    excel_path = os.path.join(output_dir, excel_filename)

    # Export to Excel
    sheets_service.export_to_excel(sheet_result['spreadsheet_id'], excel_path)

    return {
        **sheet_result,
        'excel_path': excel_path
    }


def create_excel_only_report(session_data, attendance_records, output_dir='/tmp'):
    """
    Create Excel attendance report WITHOUT Google Sheets
    Use this as a fallback when Google Sheets service account has permission issues

    Args:
        session_data: Dictionary with session information
        attendance_records: List of attendance record dictionaries
        output_dir: Directory where Excel file should be saved

    Returns:
        dict: {
            'excel_path': str,
            'title': str
        }
    """
    try:
        # Generate title
        session_name = session_data.get('session_name', 'Attendance')
        date_str = datetime.now().strftime('%Y-%m-%d')
        sheet_title = f"Attendance - {session_name} - {date_str}"

        # Generate Excel filename
        session_name_safe = session_name.replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f"Attendance_{session_name_safe}_{timestamp}.xlsx"
        excel_path = os.path.join(output_dir, excel_filename)

        # Prepare data (same format as Google Sheets)
        values = [
            ['ATTENDANCE REPORT'],
            [],
            ['Session:', session_data.get('session_name', 'N/A')],
            ['Course Code:', session_data.get('course_code', 'N/A')],
            ['Course Name:', session_data.get('course_name', 'N/A')],
            ['Instructor:', session_data.get('instructor_name', 'N/A')],
            ['Location:', session_data.get('location', 'N/A')],
            ['Start Time:', session_data.get('start_time', 'N/A')],
            ['End Time:', session_data.get('end_time', 'N/A')],
            ['Total Students:', str(len(attendance_records))],
            [],
            ['Student ID', 'Name', 'Email', 'Program', 'Time Marked', 'Status', 'Confidence Score']
        ]

        # Add attendance records
        for record in attendance_records:
            values.append([
                record.get('student_student_id', 'N/A'),
                record.get('student_name', 'N/A'),
                record.get('student_email', 'N/A'),
                record.get('student_program', 'N/A'),
                record.get('timestamp', 'N/A'),
                record.get('status', 'present').upper(),
                f"{record.get('confidence_score', 0):.4f}" if record.get('confidence_score') else 'N/A'
            ])

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Attendance'

        # Write data to Excel
        for row_idx, row in enumerate(values, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Format title row
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                    cell.alignment = Alignment(horizontal='center')

                # Format header row (row 12)
                elif row_idx == 12:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save Excel file
        wb.save(excel_path)
        print(f"[INFO] Created Excel-only report: {excel_path}")

        return {
            'excel_path': excel_path,
            'title': sheet_title
        }

    except Exception as error:
        print(f"[ERROR] Failed to create Excel report: {error}")
        raise